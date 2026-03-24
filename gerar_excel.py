"""
Gera a planilha Excel do Caixa Sublime Som
com fórmulas automáticas e gráficos bonitos.
Basta colar os dados novos na aba DADOS e tudo atualiza.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from copy import copy

# ── CORES ──
VERDE_HEX = "0D9F6E"
VERMELHO_HEX = "E02D3C"
AZUL_HEX = "1E40AF"
LARANJA_HEX = "D97706"
HEADER_BG = "0F172A"
HEADER_BG2 = "1E3A5F"
CINZA_CLARO = "F1F5F9"
CINZA_BORDA = "E2E8F0"
BRANCO = "FFFFFF"

# Fonts
font_titulo = Font(name="Calibri", size=22, bold=True, color=BRANCO)
font_subtitulo = Font(name="Calibri", size=11, color="94A3B8")
font_header = Font(name="Calibri", size=11, bold=True, color=BRANCO)
font_label = Font(name="Calibri", size=9, bold=True, color="64748B")
font_valor = Font(name="Calibri", size=18, bold=True, color="1E293B")
font_valor_verde = Font(name="Calibri", size=18, bold=True, color=VERDE_HEX)
font_valor_vermelho = Font(name="Calibri", size=18, bold=True, color=VERMELHO_HEX)
font_valor_azul = Font(name="Calibri", size=18, bold=True, color=AZUL_HEX)
font_normal = Font(name="Calibri", size=11, color="334155")
font_normal_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
font_hint = Font(name="Calibri", size=9, color="94A3B8")
font_mes_titulo = Font(name="Calibri", size=14, bold=True, color="1E293B")
font_alerta_red = Font(name="Calibri", size=11, bold=True, color=VERMELHO_HEX)
font_alerta_yel = Font(name="Calibri", size=11, bold=True, color=LARANJA_HEX)
font_alerta_grn = Font(name="Calibri", size=11, bold=True, color=VERDE_HEX)
font_instrucao = Font(name="Calibri", size=12, bold=True, color=AZUL_HEX)

# Fills
fill_header = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
fill_header2 = PatternFill(start_color=HEADER_BG2, end_color=HEADER_BG2, fill_type="solid")
fill_verde_bg = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
fill_vermelho_bg = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
fill_azul_bg = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
fill_laranja_bg = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
fill_cinza = PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type="solid")
fill_branco = PatternFill(start_color=BRANCO, end_color=BRANCO, fill_type="solid")
fill_verde_solid = PatternFill(start_color=VERDE_HEX, end_color=VERDE_HEX, fill_type="solid")
fill_vermelho_solid = PatternFill(start_color=VERMELHO_HEX, end_color=VERMELHO_HEX, fill_type="solid")
fill_azul_solid = PatternFill(start_color=AZUL_HEX, end_color=AZUL_HEX, fill_type="solid")

# Border
thin_border = Border(
    left=Side(style="thin", color=CINZA_BORDA),
    right=Side(style="thin", color=CINZA_BORDA),
    top=Side(style="thin", color=CINZA_BORDA),
    bottom=Side(style="thin", color=CINZA_BORDA),
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

BRL = '#,##0.00"";-#,##0.00""'


def estilizar_celula(ws, row, col, value=None, font=None, fill=None, alignment=None, border=None, number_format=None, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            mc = ws.cell(row=row, column=c)
            if fill: mc.fill = fill
            if border: mc.border = border
    return cell


def criar_planilha():
    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════
    # ABA 1: DADOS (onde o usuário cola dados)
    # ═══════════════════════════════════════
    ws_dados = wb.active
    ws_dados.title = "DADOS"
    ws_dados.sheet_properties.tabColor = AZUL_HEX

    # Largura das colunas
    ws_dados.column_dimensions["A"].width = 14
    ws_dados.column_dimensions["B"].width = 14
    ws_dados.column_dimensions["C"].width = 40
    ws_dados.column_dimensions["D"].width = 16
    ws_dados.column_dimensions["E"].width = 16

    # Instrução no topo
    estilizar_celula(ws_dados, 1, 1, "📋 COLE SEUS DADOS AQUI", font_instrucao, fill_azul_bg, align_left, merge_end_col=5)
    estilizar_celula(ws_dados, 2, 1, "Preencha a partir da linha 4. As colunas são: Data, Mês, Descrição, Entrada, Saída.",
                     font_hint, fill_azul_bg, align_left, merge_end_col=5)

    # Cabeçalho da tabela
    headers_dados = ["Data", "Mês", "Descrição", "Entrada (R$)", "Saída (R$)"]
    for i, h in enumerate(headers_dados, 1):
        estilizar_celula(ws_dados, 3, i, h, font_header, fill_header, align_center, thin_border)

    # Dados iniciais (Jan-Mar 2026)
    transacoes = [
        ("05/01/2026", "Janeiro", "Oferta de Ensaio", 50.32, 0),
        ("10/01/2026", "Janeiro", "Sorteio", 1410.00, 0),
        ("12/01/2026", "Janeiro", "Oferta de Ensaio", 50.00, 0),
        ("15/01/2026", "Janeiro", "Lembrancinhas (Parte)", 0, 266.15),
        ("19/01/2026", "Janeiro", "Oferta de Ensaio", 50.00, 0),
        ("28/01/2026", "Janeiro", "Compra de TNT", 0, 41.00),
        ("01/02/2026", "Fevereiro", "Oferta de Ensaio", 55.00, 0),
        ("06/02/2026", "Fevereiro", "Joum David", 0, 500.00),
        ("15/02/2026", "Fevereiro", "Oferta de Ensaio", 50.00, 0),
        ("22/02/2026", "Fevereiro", "Oferta de Ensaio", 50.00, 0),
        ("25/02/2026", "Fevereiro", "Caixa de Som JBL", 0, 50.00),
        ("01/03/2026", "Março", "Oferta de Ensaio", 50.00, 0),
        ("03/03/2026", "Março", "Lava Car", 480.34, 0),
        ("06/03/2026", "Março", "David Simpósio", 0, 320.00),
        ("08/03/2026", "Março", "Oferta de Ensaio", 50.00, 0),
        ("10/03/2026", "Março", "Gás", 0, 100.00),
        ("12/03/2026", "Março", "Decoração Dia das Mães", 0, 128.00),
        ("14/03/2026", "Março", "Sorteio Março", 450.00, 0),
        ("15/03/2026", "Março", "Oferta de Ensaio", 50.00, 0),
        ("20/03/2026", "Março", "Bolo Aniversário", 0, 100.00),
        ("22/03/2026", "Março", "Oferta de Ensaio", 50.00, 0),
        ("25/03/2026", "Março", "Lembrancinhas", 0, 168.16),
        ("28/03/2026", "Março", "Oferta de Ensaio", 50.00, 0),
        ("30/03/2026", "Março", "Uber Simpósio", 0, 138.50),
        ("31/03/2026", "Março", "Oferta de Ensaio", 50.00, 0),
    ]

    for idx, (data, mes, desc, ent, sai) in enumerate(transacoes):
        r = 4 + idx
        fill_row = fill_cinza if idx % 2 == 0 else fill_branco
        estilizar_celula(ws_dados, r, 1, data, font_normal, fill_row, align_center, thin_border)
        estilizar_celula(ws_dados, r, 2, mes, font_normal, fill_row, align_center, thin_border)
        estilizar_celula(ws_dados, r, 3, desc, font_normal, fill_row, align_left, thin_border)
        estilizar_celula(ws_dados, r, 4, ent if ent > 0 else None, font_normal, fill_row, align_right, thin_border, BRL)
        estilizar_celula(ws_dados, r, 5, sai if sai > 0 else None, font_normal, fill_row, align_right, thin_border, BRL)

    ULTIMA_LINHA_DADOS = 4 + len(transacoes) - 1
    # Definir nome para facilitar fórmulas
    MAX_DADOS = 200  # espaço para até 200 transações

    # Formatar linhas vazias futuras (até 200)
    for r in range(ULTIMA_LINHA_DADOS + 1, 4 + MAX_DADOS):
        fill_row = fill_cinza if (r - 4) % 2 == 0 else fill_branco
        for c in range(1, 6):
            cell = ws_dados.cell(row=r, column=c)
            cell.fill = fill_row
            cell.border = thin_border
            cell.font = font_normal
            if c == 4 or c == 5:
                cell.number_format = BRL
                cell.alignment = align_right
            elif c == 1 or c == 2:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # ═══════════════════════════════════════
    # ABA 2: RESUMO (calculado automaticamente)
    # ═══════════════════════════════════════
    ws = wb.create_sheet("RESUMO")
    ws.sheet_properties.tabColor = VERDE_HEX

    # Largura das colunas
    for col_letter in ["A","B","C","D","E","F","G","H","I","J","K","L"]:
        ws.column_dimensions[col_letter].width = 16
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20

    # ── HEADER ──
    for c in range(1, 9):
        estilizar_celula(ws, 1, c, fill=fill_header)
        estilizar_celula(ws, 2, c, fill=fill_header)
        estilizar_celula(ws, 3, c, fill=fill_header)

    estilizar_celula(ws, 1, 2, "🎵 CAIXA SUBLIME SOM", font_titulo, fill_header, align_left, merge_end_col=5)
    estilizar_celula(ws, 2, 2, "Relatório Financeiro — 1º Trimestre 2026", font_subtitulo, fill_header, align_left, merge_end_col=5)

    # ── FÓRMULAS DE RESUMO (usam SUMPRODUCT para filtrar por mês) ──
    # Range de dados
    DR = f"DADOS!$B$4:$B${3 + MAX_DADOS}"  # Mês
    DE = f"DADOS!$D$4:$D${3 + MAX_DADOS}"  # Entradas
    DS = f"DADOS!$E$4:$E${3 + MAX_DADOS}"  # Saídas

    meses = ["Janeiro", "Fevereiro", "Março"]
    saldo_anterior_ref = 77.34  # Saldo antes de janeiro

    # ── KPIs ── (Linha 5-8)
    row = 5
    estilizar_celula(ws, row, 2, "SALDO EM CAIXA", font_label, fill_verde_bg, align_center, thin_border)
    estilizar_celula(ws, row, 3, "TOTAL ENTRADAS", font_label, fill_verde_bg, align_center, thin_border)
    estilizar_celula(ws, row, 4, "TOTAL SAÍDAS", font_label, fill_vermelho_bg, align_center, thin_border)
    estilizar_celula(ws, row, 5, "RESULTADO", font_label, fill_azul_bg, align_center, thin_border)

    row = 6
    # Total Entradas = soma de todas entradas
    estilizar_celula(ws, row, 3, None, font_valor_verde, fill_verde_bg, align_center, thin_border, BRL)
    ws.cell(row=row, column=3).value = f"=SUM({DE})"

    # Total Saídas
    estilizar_celula(ws, row, 4, None, font_valor_vermelho, fill_vermelho_bg, align_center, thin_border, BRL)
    ws.cell(row=row, column=4).value = f"=SUM({DS})"

    # Resultado = Entradas - Saídas
    estilizar_celula(ws, row, 5, None, font_valor_azul, fill_azul_bg, align_center, thin_border, BRL)
    ws.cell(row=row, column=5).value = "=C6-D6"

    # Saldo em Caixa = Saldo anterior + Resultado
    estilizar_celula(ws, row, 2, None, font_valor_verde, fill_verde_bg, align_center, thin_border, BRL)
    ws.cell(row=row, column=2).value = f"={saldo_anterior_ref}+E6"

    row = 7
    estilizar_celula(ws, row, 2, "Dinheiro disponível", font_hint, fill_verde_bg, align_center)
    estilizar_celula(ws, row, 3, "Tudo que entrou", font_hint, fill_verde_bg, align_center)
    estilizar_celula(ws, row, 4, "Tudo que saiu", font_hint, fill_vermelho_bg, align_center)
    estilizar_celula(ws, row, 5, "Entradas - Saídas", font_hint, fill_azul_bg, align_center)

    # ── RESUMO MENSAL ── (Linha 9-18)
    row = 9
    estilizar_celula(ws, row, 2, "📅 RESUMO POR MÊS", Font(name="Calibri", size=14, bold=True, color="1E293B"), fill_branco, align_left, merge_end_col=5)

    row = 10
    for i, h in enumerate(["", "MÊS", "ENTRADAS", "SAÍDAS", "RESULTADO"], 1):
        estilizar_celula(ws, row, i, h, font_header, fill_header2, align_center, thin_border)

    cores_mes = [fill_verde_bg, fill_vermelho_bg, fill_azul_bg]
    for idx, mes in enumerate(meses):
        r = 11 + idx
        fill_m = cores_mes[idx]
        estilizar_celula(ws, r, 2, mes, font_normal_bold, fill_m, align_center, thin_border)

        # Entradas do mês = SUMPRODUCT
        estilizar_celula(ws, r, 3, None, font_normal_bold, fill_m, align_right, thin_border, BRL)
        ws.cell(row=r, column=3).value = f'=SUMPRODUCT(({DR}="{mes}")*({DE}))'

        # Saídas do mês
        estilizar_celula(ws, r, 4, None, font_normal_bold, fill_m, align_right, thin_border, BRL)
        ws.cell(row=r, column=4).value = f'=SUMPRODUCT(({DR}="{mes}")*({DS}))'

        # Resultado
        estilizar_celula(ws, r, 5, None, font_normal_bold, fill_m, align_right, thin_border, BRL)
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"

    # Total
    r_total = 14
    estilizar_celula(ws, r_total, 2, "TOTAL", Font(name="Calibri", size=11, bold=True, color=BRANCO), fill_header, align_center, thin_border)
    for c in [3, 4, 5]:
        estilizar_celula(ws, r_total, c, None, Font(name="Calibri", size=11, bold=True, color=BRANCO), fill_header, align_right, thin_border, BRL)
    ws.cell(row=r_total, column=3).value = "=SUM(C11:C13)"
    ws.cell(row=r_total, column=4).value = "=SUM(D11:D13)"
    ws.cell(row=r_total, column=5).value = "=SUM(E11:E13)"

    # ── GRÁFICO BARRAS: ENTRADAS VS SAÍDAS ──
    chart_bar = BarChart()
    chart_bar.type = "col"
    chart_bar.grouping = "clustered"
    chart_bar.title = "Entradas vs Saídas por Mês"
    chart_bar.y_axis.title = "Valor (R$)"
    chart_bar.x_axis.title = None
    chart_bar.style = 10
    chart_bar.width = 22
    chart_bar.height = 14

    cats = Reference(ws, min_col=2, min_row=11, max_row=13)
    data_ent = Reference(ws, min_col=3, min_row=10, max_row=13)
    data_sai = Reference(ws, min_col=4, min_row=10, max_row=13)

    chart_bar.add_data(data_ent, titles_from_data=True)
    chart_bar.add_data(data_sai, titles_from_data=True)
    chart_bar.set_categories(cats)

    chart_bar.series[0].graphicalProperties.solidFill = VERDE_HEX
    chart_bar.series[1].graphicalProperties.solidFill = VERMELHO_HEX

    chart_bar.series[0].dLbls = DataLabelList()
    chart_bar.series[0].dLbls.showVal = True
    chart_bar.series[0].dLbls.numFmt = '#,##0'
    chart_bar.series[1].dLbls = DataLabelList()
    chart_bar.series[1].dLbls.showVal = True
    chart_bar.series[1].dLbls.numFmt = '#,##0'

    chart_bar.legend.position = "t"
    ws.add_chart(chart_bar, "B16")

    # ── SAÍDAS POR CATEGORIA (à direita) ──
    # Categorias automáticas via fórmulas
    row_cat_start = 9
    estilizar_celula(ws, row_cat_start, 7, "📊 SAÍDAS POR CATEGORIA", Font(name="Calibri", size=14, bold=True, color="1E293B"), fill_branco, align_left, merge_end_col=8)

    categorias = [
        ("Empréstimos/Adiantamentos", ["Joum David", "David Simpósio", "Uber Simpósio"]),
        ("Operacional/Eventos", ["Gás", "Caixa de Som", "Bolo"]),
        ("Material/Decoração", ["TNT", "Decoração", "Lembrancinhas"]),
        ("Ofertas de Ensaio", ["Oferta de Ensaio"]),
    ]

    estilizar_celula(ws, 10, 7, "CATEGORIA", font_header, fill_header2, align_center, thin_border)
    estilizar_celula(ws, 10, 8, "TOTAL (R$)", font_header, fill_header2, align_center, thin_border)

    # Para cada categoria, criar SUMPRODUCT com ISNUMBER(SEARCH())
    desc_range = f"DADOS!$C$4:$C${3 + MAX_DADOS}"
    saida_range = f"DADOS!$E$4:$E${3 + MAX_DADOS}"

    cat_fills = [fill_vermelho_bg, fill_laranja_bg, fill_azul_bg, fill_verde_bg]

    for idx, (cat_nome, palavras) in enumerate(categorias):
        r = 11 + idx
        estilizar_celula(ws, r, 7, cat_nome, font_normal_bold, cat_fills[idx], align_left, thin_border)
        estilizar_celula(ws, r, 8, None, font_normal_bold, cat_fills[idx], align_right, thin_border, BRL)

        # Build SUMPRODUCT with ISNUMBER(SEARCH()) for each keyword
        parts = [f'ISNUMBER(SEARCH("{p}",{desc_range}))' for p in palavras]
        formula = "+".join(parts)
        ws.cell(row=r, column=8).value = f"=SUMPRODUCT(({formula})*({saida_range}))"

    # Outros (tudo que não encaixa)
    r_outros = 11 + len(categorias)
    estilizar_celula(ws, r_outros, 7, "Outros", font_normal_bold, fill_cinza, align_left, thin_border)
    estilizar_celula(ws, r_outros, 8, None, font_normal_bold, fill_cinza, align_right, thin_border, BRL)
    ws.cell(row=r_outros, column=8).value = f"=SUM({saida_range})-SUM(H11:H{r_outros-1})"

    # ── GRÁFICO PIZZA ──
    chart_pie = PieChart()
    chart_pie.title = "Distribuição dos Gastos"
    chart_pie.style = 10
    chart_pie.width = 18
    chart_pie.height = 14

    cats_pie = Reference(ws, min_col=7, min_row=11, max_row=r_outros)
    data_pie = Reference(ws, min_col=8, min_row=10, max_row=r_outros)
    chart_pie.add_data(data_pie, titles_from_data=True)
    chart_pie.set_categories(cats_pie)

    cores_pie = [VERMELHO_HEX, LARANJA_HEX, AZUL_HEX, VERDE_HEX, "94A3B8"]
    for i, cor in enumerate(cores_pie[:r_outros - 10]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = cor
        chart_pie.series[0].data_points.append(pt)

    chart_pie.dataLabels = DataLabelList()
    chart_pie.dataLabels.showPercent = True
    chart_pie.dataLabels.showVal = False
    chart_pie.dataLabels.showCatName = True

    ws.add_chart(chart_pie, "G16")

    # ── ALERTAS ── (Linha 32+)
    row_alert = 32
    estilizar_celula(ws, row_alert, 2, "🚨 PONTOS DE ATENÇÃO", Font(name="Calibri", size=14, bold=True, color="1E293B"), fill_branco, align_left, merge_end_col=5)

    alertas = [
        (fill_vermelho_bg, font_alerta_red, "🔴 Empréstimo pendente — David recebeu R$ 320 para simpósio e não devolveu."),
        (fill_vermelho_bg, font_alerta_red, "🔴 Sem evento = caixa negativo — Fevereiro não teve evento e ficou R$ 395 negativo."),
        (fill_laranja_bg, font_alerta_yel, "🟡 80% da receita vem de eventos — Ofertas cobrem só 20%. Diversificar receita."),
        (fill_laranja_bg, font_alerta_yel, "🟡 Gastos altos sem registro claro — Sempre anotar pra quem e pra quê."),
        (fill_verde_bg, font_alerta_grn, f"🟢 Trimestre positivo! Sobraram R$ 1.133,85. Saldo atual: R$ 786,45."),
    ]

    for idx, (fill_a, font_a, texto) in enumerate(alertas):
        r = row_alert + 1 + idx
        estilizar_celula(ws, r, 2, texto, font_a, fill_a, align_left, thin_border, merge_end_col=8)

    # ── SUGESTÕES ──
    row_sug = row_alert + 1 + len(alertas) + 1
    estilizar_celula(ws, row_sug, 2, "💡 SUGESTÕES", Font(name="Calibri", size=14, bold=True, color="1E293B"), fill_branco, align_left, merge_end_col=5)

    sugestoes = [
        "📅 Fazer pelo menos 1 evento de arrecadação por mês.",
        "📝 Anotar sempre: quem recebeu, pra quê, quando devolve.",
        "🤝 Gastos acima de R$ 200 precisam de aprovação de 2 líderes.",
        "💰 Guardar 20% de cada evento como reserva de segurança.",
        "🔄 Cobrar devoluções de empréstimos semanalmente.",
    ]

    for idx, sug in enumerate(sugestoes):
        r = row_sug + 1 + idx
        estilizar_celula(ws, r, 2, sug, font_normal, fill_verde_bg, align_left, thin_border, merge_end_col=8)

    # ── CONFIG IMPRESSÃO ──
    ws.sheet_view.showGridLines = False
    ws_dados.sheet_view.showGridLines = True

    # Definir aba ativa
    wb.active = wb.sheetnames.index("RESUMO")

    # Salvar
    caminho = "CaixaSublimeSom_Relatorio.xlsx"
    wb.save(caminho)
    print(f"✅ Planilha gerada: {caminho}")
    return caminho


if __name__ == "__main__":
    criar_planilha()
